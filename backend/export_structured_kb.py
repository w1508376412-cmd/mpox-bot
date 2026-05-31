"""Export the rebuilt knowledge base to an enhanced structured Excel workbook."""
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import psycopg
from psycopg.rows import dict_row

from config import get_settings


settings = get_settings()


EXPORT_PATH = Path("/Users/yanfei/Desktop/猴痘AI聊天机器人_结构化知识库导出_增强版.xlsx")


FIELD_DICTIONARY = [
    ("document_id", "文档唯一标识", "文档清单/知识片段关联键", "必填"),
    ("title", "文档标题", "用于人工复核和引用", "必填"),
    ("source", "归一化来源", "已将联合发布或部门名称归一到主机构", "必填"),
    ("source_primary", "主来源机构", "用于权威等级和来源统计", "推荐"),
    ("source_orgs", "联合发布机构列表", "保留多机构发布信息", "推荐"),
    ("url", "来源链接", "用于可追溯引用", "推荐"),
    ("url_status", "URL状态", "present/missing", "推荐"),
    ("publish_date", "原文发布日期", "用于时间权重排序", "必填"),
    ("date_source", "日期来源", "raw_publish_date/known_metadata/unknown", "推荐"),
    ("date_confidence", "日期可信度", "high/medium/low", "推荐"),
    ("region", "适用地区", "中国/global/欧洲/非洲等", "必填"),
    ("content_preview", "片段预览", "便于浏览检查", "推荐"),
    ("section_title", "所属章节标题", "用于判断片段上下文和来源位置", "推荐"),
    ("start_char", "片段起始字符位置", "用于回溯原文和诊断切片边界", "推荐"),
    ("end_char", "片段结束字符位置", "用于回溯原文和诊断切片边界", "推荐"),
    ("content", "片段全文", "用于 RAG 检索增强生成", "必填"),
    ("content_hash", "片段内容哈希", "用于去重和变更追踪", "推荐"),
    ("quality_score", "片段质量评分", "越高越适合进入检索", "推荐"),
    ("quality_warnings", "质量问题标签", "用于快速定位需修复内容", "推荐"),
    ("contains_mpox_keyword", "是否包含猴痘关键词", "主题过滤辅助字段", "推荐"),
    ("is_suspected_irrelevant", "是否疑似无关", "疑似混入其他疾病或非主题内容", "推荐"),
    ("embedding_model", "向量模型", "当前为阿里百炼 text-embedding-v4", "推荐"),
]


def fetch_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    with psycopg.connect(settings.database_url, row_factory=dict_row) as conn:
        documents = conn.execute(
            """
            SELECT id, title, source, source_primary, source_department, source_orgs,
                   url, url_status, url_missing_reason, publish_date, date_source,
                   date_confidence, region, original_file, processed_at, created_at, updated_at
            FROM documents
            ORDER BY source, publish_date DESC, title
            """
        ).fetchall()
        chunks = conn.execute(
            """
            SELECT id, document_id, chunk_index, section_title, start_char, end_char,
                   content_preview, content, content_hash,
                   topic, source, source_primary, source_department, url, publish_date,
                   date_confidence, region, priority, quality_score, quality_warnings,
                   contains_mpox_keyword, is_footer_like, is_suspected_irrelevant,
                   language, embedding_model, processed_at, created_at
            FROM chunks
            WHERE is_active = true
            ORDER BY source, document_id, chunk_index, id
            """
        ).fetchall()
    return list(documents), list(chunks)


def stringify(value: Any) -> Any:
    if isinstance(value, list):
        return "、".join(str(item) for item in value)
    return value


def add_sheet(wb: openpyxl.Workbook, title: str, headers: Iterable[str], rows: Iterable[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    headers = list(headers)
    ws.append(headers)
    for row in rows:
        ws.append([stringify(row.get(header, "")) for header in headers])
    style_sheet(ws)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        max_length = min(max(len(str(cell.value or "")) for cell in column), 60)
        ws.column_dimensions[letter].width = max(max_length + 2, 12)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_issue_rows(documents: list[dict[str, Any]], chunks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for doc in documents:
        if doc.get("url_status") == "missing" or not doc.get("url"):
            issues.append(
                {
                    "issue_type": "URL缺失",
                    "severity": "high",
                    "record_id": doc["id"],
                    "title": doc["title"],
                    "detail": doc.get("url_missing_reason", "缺少可追溯链接"),
                    "suggestion": "补充原文链接；本地整理材料记录原始文件来源",
                }
            )
        if doc.get("date_confidence") == "low":
            issues.append(
                {
                    "issue_type": "发布日期可信度低",
                    "severity": "high",
                    "record_id": doc["id"],
                    "title": doc["title"],
                    "detail": f"publish_date={doc.get('publish_date')}, date_source={doc.get('date_source')}",
                    "suggestion": "人工核对原文发布日期，不要用处理日期替代发布日期",
                }
            )
    for chunk in chunks:
        warnings = chunk.get("quality_warnings") or []
        if chunk.get("quality_score", 1) < 0.8 or warnings:
            issues.append(
                {
                    "issue_type": "片段质量需复核",
                    "severity": "medium" if chunk.get("quality_score", 1) >= 0.5 else "high",
                    "record_id": chunk["id"],
                    "title": chunk["document_id"],
                    "detail": "、".join(warnings) if isinstance(warnings, list) else str(warnings),
                    "suggestion": "检查是否为页脚、过短片段、无关主题或切片边界问题",
                }
            )
    return issues


def build_summary_rows(documents: list[dict[str, Any]], chunks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    source_counts = Counter(chunk["source"] for chunk in chunks)
    low_quality_count = sum(1 for chunk in chunks if chunk.get("quality_score", 1) < 0.8)
    missing_url_count = sum(1 for doc in documents if doc.get("url_status") == "missing" or not doc.get("url"))
    low_date_count = sum(1 for doc in documents if doc.get("date_confidence") == "low")
    rows = [
        {"指标": "导出时间", "值": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"指标": "文档数", "值": len(documents)},
        {"指标": "知识片段数", "值": len(chunks)},
        {"指标": "低质量片段数(quality_score < 0.8)", "值": low_quality_count},
        {"指标": "URL缺失文档数", "值": missing_url_count},
        {"指标": "发布日期可信度低文档数", "值": low_date_count},
        {"指标": "向量配置", "值": "pgvector embedding vector(1024) + embedding_json"},
    ]
    for source, count in source_counts.most_common():
        rows.append({"指标": f"来源统计 - {source}", "值": count})
    return rows


def export_workbook(output_path: Path = EXPORT_PATH) -> Path:
    documents, chunks = fetch_rows()
    issue_rows = build_issue_rows(documents, chunks)
    summary_rows = build_summary_rows(documents, chunks)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "导出说明", ["指标", "值"], summary_rows)
    add_sheet(wb, "问题清单", ["issue_type", "severity", "record_id", "title", "detail", "suggestion"], issue_rows)
    add_sheet(wb, "文档清单", list(documents[0].keys()) if documents else [], documents)
    chunk_headers = [
        "id",
        "document_id",
        "chunk_index",
        "section_title",
        "start_char",
        "end_char",
        "content_preview",
        "content",
        "content_hash",
        "topic",
        "source",
        "source_primary",
        "source_department",
        "url",
        "publish_date",
        "date_confidence",
        "region",
        "priority",
        "quality_score",
        "quality_warnings",
        "contains_mpox_keyword",
        "is_footer_like",
        "is_suspected_irrelevant",
        "language",
        "embedding_model",
        "processed_at",
        "created_at",
    ]
    add_sheet(wb, "知识片段", chunk_headers, chunks)
    add_sheet(wb, "字段字典", ["字段名", "中文说明", "用途", "建议"], [
        {"字段名": name, "中文说明": description, "用途": purpose, "建议": requirement}
        for name, description, purpose, requirement in FIELD_DICTIONARY
    ])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = export_workbook()
    print(path)
